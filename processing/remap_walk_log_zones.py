from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ZONE_LABEL_RE = re.compile(r"^(Inwood|WaHi)\s+(\d+)\s*$", re.IGNORECASE)

# Pastel palette for Excel column A zone rows (Zone_ID 1..25).
EXCEL_ZONE_COLORS: dict[int, str] = {
    1: "#bfe5ba",
    2: "#b7dee3",
    3: "#afdfc3",
    4: "#cfbade",
    5: "#ead2c6",
    6: "#f5f5c8",
    7: "#b9e8cc",
    8: "#e7c0ce",
    9: "#c2cbed",
    10: "#bfe5ba",
    11: "#ead2c6",
    12: "#f5f5c8",
    13: "#afdfc3",
    14: "#b7dee3",
    15: "#e7c0ce",
    16: "#cfbade",
    17: "#ead2c6",
    18: "#bdd1e8",
    19: "#bfe5ba",
    20: "#ead2c6",
    21: "#c2cbed",
    22: "#ead2c6",
    23: "#f5f5c8",
    24: "#b7dee3",
    25: "#e7c0ce",
}


def has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def old_zone_id_from_label(label: Any) -> str | None:
    text = "" if label is None else str(label).strip()
    match = ZONE_LABEL_RE.fullmatch(text)
    if not match:
        return None

    prefix = match.group(1).lower()
    zone_num = int(match.group(2))
    if prefix == "inwood":
        return f"I_{zone_num}"
    return f"WH_{zone_num}"


def load_zone_mapping(zones_geojson_path: Path) -> dict[str, int]:
    with zones_geojson_path.open("r", encoding="utf-8") as f:
        geojson_obj = json.load(f)

    mapping: dict[str, int] = {}
    for feature in geojson_obj.get("features", []):
        if not isinstance(feature, dict):
            continue

        properties = feature.get("properties")
        if not isinstance(properties, dict):
            continue

        old_zone_id = properties.get("Old_Zone_ID")
        new_zone_id = properties.get("Zone_ID")
        if old_zone_id is None or new_zone_id is None:
            continue

        old_zone_id_text = str(old_zone_id).strip()
        if not old_zone_id_text:
            continue

        try:
            new_zone_num = int(new_zone_id)
        except (TypeError, ValueError):
            continue

        mapping[old_zone_id_text] = new_zone_num

    if len(mapping) < 25:
        raise ValueError(
            f"Expected at least 25 zone mappings from {zones_geojson_path}, found {len(mapping)}"
        )

    return mapping


def process_sheet(sheet: Any, zone_mapping: dict[str, int]) -> tuple[int, int]:
    # Requested workbook cleanup.
    sheet.delete_rows(18, 2)

    zone_rows: list[tuple[int, str]] = []
    for row_idx in range(1, (sheet.max_row or 0) + 1):
        old_zone_id = old_zone_id_from_label(sheet.cell(row=row_idx, column=1).value)
        if old_zone_id is not None:
            zone_rows.append((row_idx, old_zone_id))

    if len(zone_rows) != 25:
        raise ValueError(
            f"Sheet {sheet.title!r}: expected 25 zone rows after deleting rows 18/19, found {len(zone_rows)}"
        )

    zone_rows.sort(key=lambda item: item[0])
    ordered_zone_rows = [row_idx for row_idx, _ in zone_rows]
    source_row_by_old_id = {old_zone_id: row_idx for row_idx, old_zone_id in zone_rows}

    target_row_by_new_zone = {
        zone_num: row_idx for zone_num, row_idx in enumerate(ordered_zone_rows, start=1)
    }

    max_col = sheet.max_column or 1
    moved_values: dict[tuple[int, int], Any] = {}

    for old_zone_id, source_row in source_row_by_old_id.items():
        if old_zone_id not in zone_mapping:
            raise ValueError(f"Sheet {sheet.title!r}: no mapping found for old zone id {old_zone_id!r}")

        new_zone_num = zone_mapping[old_zone_id]
        if new_zone_num not in target_row_by_new_zone:
            raise ValueError(
                f"Sheet {sheet.title!r}: mapped new zone number out of range: {new_zone_num}"
            )

        target_row = target_row_by_new_zone[new_zone_num]

        for col_idx in range(2, max_col + 1):
            value = sheet.cell(row=source_row, column=col_idx).value
            if not has_value(value):
                continue

            target_key = (target_row, col_idx)
            if target_key in moved_values:
                raise ValueError(
                    f"Sheet {sheet.title!r}: collision while moving values into row {target_row}, col {col_idx}"
                )

            moved_values[target_key] = value

    # Clear only zone rows, only non-label columns.
    for row_idx in ordered_zone_rows:
        for col_idx in range(2, max_col + 1):
            sheet.cell(row=row_idx, column=col_idx).value = None

    # Write moved values to their new rows, preserving original columns.
    for (row_idx, col_idx), value in moved_values.items():
        sheet.cell(row=row_idx, column=col_idx).value = value

    # Replace column A zone labels with 1..25 from top to bottom,
    # and apply the requested Excel fill colors.
    for zone_num, row_idx in target_row_by_new_zone.items():
        cell = sheet.cell(row=row_idx, column=1)
        cell.value = zone_num

        zone_hex = EXCEL_ZONE_COLORS.get(zone_num)
        if zone_hex:
            cell.fill = make_fill(zone_hex)

    return len(ordered_zone_rows), len(moved_values)


def default_output_path(input_workbook: Path) -> Path:
    return input_workbook.with_name(f"{input_workbook.stem}_edited{input_workbook.suffix}")


def to_argb(hex_color: str) -> str:
    clean = str(hex_color or "").strip().lstrip("#")
    if len(clean) != 6:
        raise ValueError(f"Expected #RRGGBB color, got: {hex_color!r}")
    return f"FF{clean.upper()}"


def make_fill(hex_color: str) -> PatternFill:
    argb = to_argb(hex_color)
    return PatternFill(fill_type="solid", start_color=argb, end_color=argb)


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]

    parser = argparse.ArgumentParser(
        description="Remap zone rows in 2026 walk-log workbook using zones.geojson mapping."
    )
    parser.add_argument(
        "--input-workbook",
        default=str(repo_root / "processing" / "2026-log-of-walks-inwood_wahi.xlsx"),
        help="Path to source XLSX file.",
    )
    parser.add_argument(
        "--zones-geojson",
        default=str(repo_root / "data" / "zones.geojson"),
        help="Path to zones GeoJSON containing Old_Zone_ID and new integer Zone_ID.",
    )
    parser.add_argument(
        "--output-workbook",
        default=None,
        help="Path to output XLSX. Default appends _edited to input workbook name.",
    )
    args = parser.parse_args()

    input_workbook = Path(args.input_workbook)
    zones_geojson = Path(args.zones_geojson)
    output_workbook = Path(args.output_workbook) if args.output_workbook else default_output_path(input_workbook)

    zone_mapping = load_zone_mapping(zones_geojson)

    workbook = load_workbook(input_workbook)
    sheets_processed = len(workbook.worksheets)
    try:
        total_cells_moved = 0
        for sheet in workbook.worksheets:
            _, moved_cells = process_sheet(sheet, zone_mapping)
            total_cells_moved += moved_cells

        output_workbook.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_workbook)
    finally:
        workbook.close()

    print(f"Saved edited workbook: {output_workbook}")
    print(f"Sheets processed: {sheets_processed}")
    print(f"Cells moved across all sheets: {total_cells_moved}")


if __name__ == "__main__":
    main()
