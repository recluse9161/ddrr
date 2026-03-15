from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WEEKDAY_CODES = ["M", "T", "W", "Th", "F", "Sa", "S"]


def load_old_to_new_zone_mapping(zones_geojson_obj: dict[str, Any]) -> dict[str, str]:
    mapping: dict[str, str] = {}

    for feature in zones_geojson_obj.get("features", []):
        if not isinstance(feature, dict):
            continue

        props = feature.get("properties")
        if not isinstance(props, dict):
            continue

        old_zone_id = props.get("Old_Zone_ID")
        new_zone_id = props.get("Zone_ID")
        if old_zone_id is None or new_zone_id is None:
            continue

        old_text = str(old_zone_id).strip()
        if not old_text:
            continue

        try:
            new_zone_num = int(new_zone_id)
        except (TypeError, ValueError):
            continue

        mapping[old_text] = str(new_zone_num)

    return mapping


def get_zone_id(label: Any, old_to_new_zone_mapping: dict[str, str] | None = None) -> str | None:
    if label is None:
        return None

    if isinstance(label, (int, float)) and not isinstance(label, bool):
        as_int = int(label)
        if float(label) == float(as_int) and 1 <= as_int <= 25:
            return str(as_int)

    clean = str(label).strip()
    if not clean:
        return None

    numeric_match = re.fullmatch(r"\d+", clean)
    if numeric_match:
        zone_num = int(clean)
        if 1 <= zone_num <= 25:
            return str(zone_num)

    inwood_match = re.fullmatch(r"(?i)Inwood\s+(\d+)\s*", clean)
    if inwood_match:
        old_zone_id = f"I_{int(inwood_match.group(1))}"
        if old_to_new_zone_mapping and old_zone_id in old_to_new_zone_mapping:
            return old_to_new_zone_mapping[old_zone_id]
        return old_zone_id

    wahi_match = re.fullmatch(r"(?i)WaHi\s+(\d+)\s*", clean)
    if wahi_match:
        old_zone_id = f"WH_{int(wahi_match.group(1))}"
        if old_to_new_zone_mapping and old_zone_id in old_to_new_zone_mapping:
            return old_to_new_zone_mapping[old_zone_id]
        return old_zone_id

    if re.fullmatch(r"(?i)Dispatch\s*", clean):
        return "Dispatch"

    return None


def parse_week_date(text_value: Any, default_year: int) -> date:
    text = "" if text_value is None else str(text_value).strip()
    if not text:
        raise ValueError("Could not parse week date from empty text value.")

    # Examples:
    # - Wk of 2/23
    # - Wk of 1/19/2026
    # - Mon, 02/23
    match = re.search(
        r"(?i)(?:Wk\s*of\s*)?(\d{1,2})\s*/\s*(\d{1,2})(?:\s*/\s*(\d{2,4}))?",
        text,
    )
    if not match:
        month_name_match = re.search(
            r"(?i)\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\b\s*(\d{1,2})",
            text,
        )
        if not month_name_match:
            raise ValueError(f"Could not parse week date from text value: {text!r}")

        month_lookup = {
            "jan": 1,
            "feb": 2,
            "mar": 3,
            "apr": 4,
            "may": 5,
            "jun": 6,
            "jul": 7,
            "aug": 8,
            "sep": 9,
            "sept": 9,
            "oct": 10,
            "nov": 11,
            "dec": 12,
        }
        month = month_lookup[month_name_match.group(1).lower()]
        day = int(month_name_match.group(2))
        year = default_year
        return date(year, month, day)

    month = int(match.group(1))
    day = int(match.group(2))

    year_text = match.group(3)
    if year_text:
        year = int(year_text)
        if len(year_text) == 2:
            year += 2000
    else:
        year = default_year

    return date(year, month, day)


def resolve_sheet_week_date(sheet: Any, default_year: int) -> date:
    candidate_texts: list[str] = []

    def add_candidate(value: Any) -> None:
        text = "" if value is None else str(value).strip()
        if text:
            candidate_texts.append(text)

    add_candidate(sheet["A1"].value)

    max_col = sheet.max_column or 0
    for col in range(1, max_col + 1):
        add_candidate(sheet.cell(row=1, column=col).value)

    add_candidate(getattr(sheet, "title", ""))

    seen: set[str] = set()
    unique_candidates: list[str] = []
    for text in candidate_texts:
        if text in seen:
            continue
        seen.add(text)
        unique_candidates.append(text)

    errors: list[str] = []
    for text in unique_candidates:
        try:
            return parse_week_date(text, default_year)
        except ValueError as exc:
            errors.append(str(exc))

    raise ValueError(
        "Could not parse week date for sheet "
        f"{getattr(sheet, 'title', '<unknown>')!r}. Checked: {unique_candidates!r}"
    )


def zone_sort_key(zone_id: str) -> tuple[int, int, str]:
    numeric_match = re.fullmatch(r"\d+", str(zone_id).strip())
    if numeric_match:
        return (0, int(zone_id), "")

    inwood_match = re.fullmatch(r"I_(\d+)", zone_id)
    if inwood_match:
        return (1, int(inwood_match.group(1)), "")

    wahi_match = re.fullmatch(r"WH_(\d+)", zone_id)
    if wahi_match:
        return (2, int(wahi_match.group(1)), "")

    if zone_id == "Dispatch":
        return (3, 999, "")

    return (9, 999, zone_id)


def has_value(cell_value: Any) -> bool:
    if cell_value is None:
        return False
    return str(cell_value).strip() != ""


def to_int_or_float(value: Any) -> int | float:
    if isinstance(value, (int, float)):
        return value
    text = "" if value is None else str(value).strip()
    if not text:
        return 0
    try:
        as_float = float(text)
    except ValueError:
        return 0

    if as_float.is_integer():
        return int(as_float)
    return as_float


def parse_day_code_from_header(value: Any) -> str | None:
    text = "" if value is None else str(value).strip().lower()
    if not text:
        return None

    if "mon" in text:
        return "M"
    if "tue" in text:
        return "T"
    if "wed" in text:
        return "W"
    if "thu" in text:
        return "Th"
    if "fri" in text:
        return "F"
    if "sat" in text:
        return "Sa"
    if "sun" in text:
        return "S"

    return None


def get_day_period_columns(sheet: Any) -> list[tuple[int, str, str]]:
    max_col = sheet.max_column or 0
    columns: list[tuple[int, str, str]] = []
    fallback_index = 0

    for col in range(2, max_col + 1):
        period = str(sheet.cell(row=2, column=col).value or "").strip().upper()
        if period not in {"AM", "PM"}:
            continue

        day_code = parse_day_code_from_header(sheet.cell(row=1, column=col).value)
        if day_code is None and col > 2:
            day_code = parse_day_code_from_header(sheet.cell(row=1, column=col - 1).value)

        if day_code is None:
            day_code = WEEKDAY_CODES[min(fallback_index // 2, len(WEEKDAY_CODES) - 1)]

        fallback_index += 1
        columns.append((col, day_code, period))

    return columns


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def is_point_on_segment(
    px: float,
    py: float,
    x1: float,
    y1: float,
    x2: float,
    y2: float,
    eps: float = 1e-9,
) -> bool:
    cross = (py - y1) * (x2 - x1) - (px - x1) * (y2 - y1)
    if abs(cross) > eps:
        return False

    min_x = min(x1, x2) - eps
    max_x = max(x1, x2) + eps
    min_y = min(y1, y2) - eps
    max_y = max(y1, y2) + eps
    return min_x <= px <= max_x and min_y <= py <= max_y


def point_in_ring(point: tuple[float, float], ring: list[Any]) -> bool:
    if not isinstance(ring, list) or len(ring) < 4:
        return False

    px, py = point
    inside = False

    for i in range(len(ring) - 1):
        p1 = ring[i]
        p2 = ring[i + 1]
        if (
            not isinstance(p1, list)
            or not isinstance(p2, list)
            or len(p1) < 2
            or len(p2) < 2
            or not is_number(p1[0])
            or not is_number(p1[1])
            or not is_number(p2[0])
            or not is_number(p2[1])
        ):
            continue

        x1, y1 = float(p1[0]), float(p1[1])
        x2, y2 = float(p2[0]), float(p2[1])

        if is_point_on_segment(px, py, x1, y1, x2, y2):
            return True

        intersects = ((y1 > py) != (y2 > py)) and (
            px < ((x2 - x1) * (py - y1) / ((y2 - y1) if (y2 - y1) != 0 else 1e-12) + x1)
        )
        if intersects:
            inside = not inside

    return inside


def point_in_polygon(point: tuple[float, float], polygon_coords: list[Any]) -> bool:
    if not isinstance(polygon_coords, list) or not polygon_coords:
        return False

    outer_ring = polygon_coords[0]
    if not point_in_ring(point, outer_ring):
        return False

    # Any hole containing the point means point is outside polygon area.
    for hole in polygon_coords[1:]:
        if point_in_ring(point, hole):
            return False

    return True


def point_in_geometry(point: tuple[float, float], geometry: dict[str, Any] | None) -> bool:
    if not isinstance(geometry, dict):
        return False

    geom_type = geometry.get("type")
    coords = geometry.get("coordinates")

    if geom_type == "Polygon":
        return point_in_polygon(point, coords)

    if geom_type == "MultiPolygon" and isinstance(coords, list):
        return any(point_in_polygon(point, polygon) for polygon in coords)

    return False


def extract_point_features(geojson_obj: dict[str, Any]) -> list[tuple[float, float]]:
    points: list[tuple[float, float]] = []
    for feature in geojson_obj.get("features", []):
        if not isinstance(feature, dict):
            continue

        geometry = feature.get("geometry")
        if not isinstance(geometry, dict):
            continue

        if geometry.get("type") != "Point":
            continue

        coords = geometry.get("coordinates")
        if (
            not isinstance(coords, list)
            or len(coords) < 2
            or not is_number(coords[0])
            or not is_number(coords[1])
        ):
            continue

        points.append((float(coords[0]), float(coords[1])))

    return points


def get_record_value(record: dict[str, Any], candidate_fields: list[str]) -> Any:
    for field_name in candidate_fields:
        if field_name in record:
            return record[field_name]

        for key, value in record.items():
            if str(key).strip().lower() == str(field_name).strip().lower():
                return value

    return None


def parse_coordinate(value: Any) -> float | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def extract_point_features_from_csv(csv_path: Path) -> list[tuple[float, float]]:
    points: list[tuple[float, float]] = []

    with csv_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for record in reader:
            longitude = parse_coordinate(
                get_record_value(record, ["Longitude", "longitude", "lon", "lng"])
            )
            latitude = parse_coordinate(
                get_record_value(record, ["Latitude", "latitude", "lat"])
            )

            if longitude is None or latitude is None:
                continue

            points.append((longitude, latitude))

    return points


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]

    parser = argparse.ArgumentParser(description="Generate walk_log.csv and joined walk-log GeoJSON outputs.")
    parser.add_argument(
        "--input-workbook",
        default=str(repo_root / "processing" / "2026-log-of-walks-inwood_wahi.xlsx"),
    )
    parser.add_argument(
        "--output-csv",
        default=str(repo_root / "processing" / "walk_log.csv"),
    )
    parser.add_argument(
        "--input-zones-geojson",
        default=str(repo_root / "data" / "zones.geojson"),
    )
    parser.add_argument(
        "--output-zones-geojson",
        default=str(repo_root / "data" / "zones_processed.geojson"),
    )
    parser.add_argument(
        "--input-sightings-csv",
        default=str(repo_root / "processing" / "sightings.csv"),
    )
    parser.add_argument(
        "--input-sightings-geojson",
        default=str(repo_root / "data" / "confirmed_sightings.geojson"),
    )
    parser.add_argument(
        "--output-dispatch-geojson",
        default=str(repo_root / "data" / "dispatch_walk_log.geojson"),
    )
    parser.add_argument("--default-year", type=int, default=2026)
    args = parser.parse_args()

    input_workbook = Path(args.input_workbook)
    output_csv = Path(args.output_csv)
    input_zones_geojson = Path(args.input_zones_geojson)
    output_zones_geojson = Path(args.output_zones_geojson)
    input_sightings_csv = Path(args.input_sightings_csv)
    input_sightings_geojson = Path(args.input_sightings_geojson)
    output_dispatch_geojson = Path(args.output_dispatch_geojson)

    with input_zones_geojson.open("r", encoding="utf-8") as f:
        zones_geojson = json.load(f)

    old_to_new_zone_mapping = load_old_to_new_zone_mapping(zones_geojson)

    workbook = load_workbook(input_workbook, data_only=True)

    weeks: list[dict[str, Any]] = []
    all_zones: set[str] = set()

    for sheet in workbook.worksheets:
        week_date = resolve_sheet_week_date(sheet, args.default_year)
        week_label = f"{week_date.month}/{week_date.day}/{week_date.year % 100:02d}"

        max_row = sheet.max_row or 0
        day_period_columns = get_day_period_columns(sheet)

        counts_by_zone: dict[str, dict[str, dict[str, int]]] = {}

        for row in range(1, max_row + 1):
            raw_label = str(sheet.cell(row=row, column=1).value or "").strip()
            if not raw_label:
                continue
            if re.fullmatch(r"(?i)Other\s*", raw_label):
                continue

            zone_id = get_zone_id(raw_label, old_to_new_zone_mapping)
            if not zone_id:
                continue

            all_zones.add(zone_id)
            zone_counts = counts_by_zone.setdefault(
                zone_id,
                {
                    day_code: {"AM": 0, "PM": 0}
                    for day_code in WEEKDAY_CODES
                },
            )

            for col, day_code, period in day_period_columns:
                if has_value(sheet.cell(row=row, column=col).value):
                    zone_counts[day_code][period] += 1

        weeks.append(
            {
                "week_date": week_date,
                "week_label": week_label,
                "week_sort": week_date.year * 10000 + week_date.month * 100 + week_date.day,
                "counts_by_zone": counts_by_zone,
            }
        )

    workbook.close()

    all_zones.add("Dispatch")
    ordered_weeks = sorted(weeks, key=lambda w: w["week_date"])
    ordered_zones = sorted(all_zones, key=zone_sort_key)

    rows: list[dict[str, Any]] = []
    for zone_id in ordered_zones:
        for week in ordered_weeks:
            zone_counts = week["counts_by_zone"].get(zone_id, {})

            for day_code in WEEKDAY_CODES:
                day_counts = zone_counts.get(day_code, {"AM": 0, "PM": 0})
                rows.append(
                    {
                        "Zone_ID": zone_id,
                        "Week_Label": week["week_label"],
                        "Week_Sort": int(week["week_sort"]),
                        "Week_Start": week["week_date"].isoformat(),
                        "Day_Code": day_code,
                        "AM_Count": int(day_counts.get("AM", 0)),
                        "PM_Count": int(day_counts.get("PM", 0)),
                    }
                )

    fieldnames: list[str] = [
        "Zone_ID",
        "Week_Label",
        "Week_Sort",
        "Week_Start",
        "Day_Code",
        "AM_Count",
        "PM_Count",
    ]

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Created CSV: {output_csv}")

    if input_sightings_csv.exists():
        sighting_points = extract_point_features_from_csv(input_sightings_csv)
    elif input_sightings_geojson.exists():
        with input_sightings_geojson.open("r", encoding="utf-8") as f:
            sightings_geojson = json.load(f)
        sighting_points = extract_point_features(sightings_geojson)
    else:
        raise FileNotFoundError(
            "Could not find sightings input. Checked CSV and GeoJSON: "
            f"{input_sightings_csv} ; {input_sightings_geojson}"
        )

    for feature in zones_geojson.get("features", []):
        props = feature.get("properties")
        if not isinstance(props, dict):
            continue

        # Count confirmed ICE sighting points inside this zone polygon/multipolygon.
        zone_geometry = feature.get("geometry")
        sighting_count = sum(
            1 for point in sighting_points if point_in_geometry(point, zone_geometry)
        )
        props["# of ICE sightings"] = int(sighting_count)

    output_zones_geojson.parent.mkdir(parents=True, exist_ok=True)
    with output_zones_geojson.open("w", encoding="utf-8") as f:
        json.dump(zones_geojson, f, ensure_ascii=False)
    print(f"Created joined zones GeoJSON: {output_zones_geojson}")

    dispatch_props: dict[str, Any] = {"Zone_ID": "Dispatch"}

    dispatch_collection: dict[str, Any] = {
        "type": "FeatureCollection",
        "name": "dispatch_walk_log",
        "features": [
            {
                "type": "Feature",
                "properties": {
                    **dispatch_props,
                    "Zone_ID": str(dispatch_props.get("Zone_ID", "Dispatch")),
                },
                "geometry": {
                    "type": "Point",
                    # GeoJSON uses [longitude, latitude]
                    "coordinates": [-73.95003953, 40.85578134],
                },
            }
        ],
    }

    if "crs" in zones_geojson:
        dispatch_collection["crs"] = zones_geojson["crs"]

    output_dispatch_geojson.parent.mkdir(parents=True, exist_ok=True)
    with output_dispatch_geojson.open("w", encoding="utf-8") as f:
        json.dump(dispatch_collection, f, ensure_ascii=False)
    print(f"Created dispatch point GeoJSON: {output_dispatch_geojson}")


if __name__ == "__main__":
    main()
